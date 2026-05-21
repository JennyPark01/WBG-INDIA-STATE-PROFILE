"""
Extract revenue data (tab 2_1) from India PFR Tool 1.xlsx
Produces pfr_2_1_data.csv in the same format as pfr_1_1_data.csv:
  State, Code, Indicator, Year, Value
"""
import openpyxl, csv, warnings
warnings.filterwarnings("ignore")

XLSX = "C:/Users/jaeye/Downloads/State profile/India PFR Tool 1.xlsx"
OUT  = "C:/Users/jaeye/Downloads/State profile/pfr_2_1_data.csv"

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 1. State names from _States sheet ──
ws_states = wb["_States"]
state_names = []
for i, row in enumerate(ws_states.iter_rows(min_col=1, max_col=1, values_only=True)):
    if i == 0:
        continue
    name = row[0]
    if name:
        state_names.append(str(name).strip())
    else:
        break
print(f"States: {len(state_names)}")

# ── 2. Year columns from _Data header ──
ws_data = wb["_Data"]
year_indices = {}  # col_idx (1-based) -> year
for cell in next(ws_data.iter_rows(min_row=1, max_row=1)):
    v = cell.value
    if v is not None:
        try:
            yr = int(float(str(v)))
            if 1989 <= yr <= 2024:
                year_indices[cell.column] = yr
        except (ValueError, TypeError):
            pass
years = sorted(y for y in year_indices.values() if y >= 1990)
print(f"Years: {years[0]}-{years[-1]} ({len(years)} years)")

# ── 3. Raw variable codes to extract ──
raw_codes = [
    "NGSDP",
    "REV_TOTAL",
    "REV_TAX_I",
    "REV_TAX_I_A",
    "REV_TAX_I_A_1",
    "REV_TAX_I_A_2",
    "REV_TAX_I_A_3",
    "REV_TAX_I_B",
    "REV_NTAX_II",
    "REV_NTAX_II_C",
    "REV_NTAX_II_D",
    "CAP_REC_III",
    "CAP_REC_XI",
]

# Build lookup: key_pattern -> set of (state, code) we're looking for
target_keys = {}
for state in state_names:
    for code in raw_codes:
        key = state + code + "None"
        target_keys[key] = (state, code)

# ── 4. Read raw data from _Data ──
raw_data = {s: {} for s in state_names}
for row in ws_data.iter_rows(min_row=2, values_only=False):
    key_cell = row[0].value
    if not key_cell:
        continue
    key = str(key_cell)
    if key in target_keys:
        state, code = target_keys[key]
        raw_data[state][code] = {}
        for col_idx, yr in year_indices.items():
            cell = row[col_idx - 1]  # 0-based in tuple
            if cell.value is not None:
                try:
                    raw_data[state][code][yr] = float(cell.value)
                except (ValueError, TypeError):
                    pass

# Verify extraction
found = sum(1 for s in state_names for c in raw_codes if c in raw_data[s])
print(f"Extracted {found} state-code combinations (expected {len(state_names) * len(raw_codes)})")

# ── 5. Compute derived metrics ──
output_rows = []

# Revenue codes (excluding NGSDP) and their display labels + output codes
rev_indicators = [
    ("REV_TOTAL",      "cur_rev",       "Current revenues"),
    ("REV_TAX_I",      "tax_rev",       "Tax revenues"),
    ("REV_TAX_I_A",    "own_tax",       "State's own tax revenues"),
    ("REV_TAX_I_A_1",  "inc_tax",       "Taxes on Income"),
    ("REV_TAX_I_A_2",  "prop_tax",      "Taxes on Property and Capital"),
    ("REV_TAX_I_A_3",  "comm_tax",      "Taxes on Commodities and Services"),
    ("REV_TAX_I_B",    "central_share", "Share in central government tax"),
    ("REV_NTAX_II",    "nontax",        "Non-tax revenues"),
    ("REV_NTAX_II_C",  "own_nontax",    "State's Own Non-Tax Revenue"),
    ("REV_NTAX_II_D",  "grants",        "Grants from the Centre"),
    ("CAP_REC_III",    "loan_rec",      "Recovery of loans and advances"),
    ("CAP_REC_XI",     "misc_cap",      "Miscellaneous capital receipts"),
]

for state in state_names:
    d = raw_data[state]
    for yr in years:
        ngdp = d.get("NGSDP", {}).get(yr)
        if not ngdp or ngdp == 0:
            continue

        # Total revenues = REV_TOTAL + CAP_REC_III + CAP_REC_XI
        rev_total = d.get("REV_TOTAL", {}).get(yr, 0)
        cap_iii   = d.get("CAP_REC_III", {}).get(yr, 0)
        cap_xi    = d.get("CAP_REC_XI", {}).get(yr, 0)
        total_rev = rev_total + cap_iii + cap_xi

        # Total revenues % of GDP
        val = total_rev / ngdp * 100
        output_rows.append([state, "total_rev_gdp", "Total revenues", yr, round(val, 2)])

        # Each revenue component as % of GDP
        for raw_code, out_code, label in rev_indicators:
            raw_val = d.get(raw_code, {}).get(yr)
            if raw_val is not None:
                pct_gdp = raw_val / ngdp * 100
                output_rows.append([state, f"{out_code}_gdp", label, yr, round(pct_gdp, 2)])

        # Percent of total revenues (only if total_rev > 0)
        if total_rev > 0:
            # Current revenues as % of total
            if rev_total:
                output_rows.append([state, "cur_rev_total", "Current revenues", yr,
                                    round(rev_total / total_rev * 100, 2)])

            for raw_code, out_code, label in rev_indicators:
                if out_code == "cur_rev":
                    continue  # already handled above
                raw_val = d.get(raw_code, {}).get(yr)
                if raw_val is not None:
                    pct_total = raw_val / total_rev * 100
                    output_rows.append([state, f"{out_code}_total", label, yr,
                                        round(pct_total, 2)])

        # Tax revenue sub-breakdown as % of total tax revenue
        tax_rev = d.get("REV_TAX_I", {}).get(yr)
        if tax_rev and tax_rev > 0:
            tax_sub_indicators = [
                ("REV_TAX_I_A_1",  "inc_tax_taxrev",   "Taxes on Income"),
                ("REV_TAX_I_A_2",  "prop_tax_taxrev",  "Taxes on Property and Capital"),
                ("REV_NTAX_II",    "nontax_taxrev",    "Non-tax revenues"),
                ("REV_NTAX_II_C",  "own_nontax_taxrev","State's Own Non-Tax Revenue"),
            ]
            for raw_code, out_code, label in tax_sub_indicators:
                raw_val = d.get(raw_code, {}).get(yr)
                if raw_val is not None:
                    pct_tax = raw_val / tax_rev * 100
                    output_rows.append([state, out_code, label, yr, round(pct_tax, 2)])

# ── 6. Write CSV ──
with open(OUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["State", "Code", "Indicator", "Year", "Value"])
    writer.writerows(output_rows)

print(f"\nWritten {len(output_rows)} rows to {OUT}")

# Verify a sample
for r in output_rows:
    if r[0] == "Bihar" and r[1] == "total_rev_gdp" and r[3] == 2017:
        print(f"Bihar total_rev_gdp 2017: {r[4]}")
    if r[0] == "Bihar" and r[1] == "tax_rev_gdp" and r[3] == 2017:
        print(f"Bihar tax_rev_gdp 2017: {r[4]}")

from collections import Counter
code_counts = Counter(r[1] for r in output_rows)
print(f"\nTotal indicator codes: {len(code_counts)}")
for code, count in sorted(code_counts.items()):
    print(f"  {code}: {count} rows")

wb.close()
