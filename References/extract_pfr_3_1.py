"""
Extract expenditure data (tab 3_1) from India PFR Tool 1.xlsx
Produces pfr_3_1_data.csv in the same format as pfr_1_1_data.csv:
  State, Code, Indicator, Year, Value
"""
import openpyxl, csv, warnings
warnings.filterwarnings("ignore")

XLSX = "C:/Users/jaeye/Downloads/State profile/India PFR Tool 1.xlsx"
OUT  = "C:/Users/jaeye/Downloads/State profile/pfr_3_1_data.csv"

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── 1. State names from _States sheet ──
ws_states = wb["_States"]
state_names = []
for i, row in enumerate(ws_states.iter_rows(min_col=1, max_col=1, values_only=True)):
    if i == 0:
        continue
    name = row[0]
    if name:
        state_names.append(str(name).strip())
    else:
        break
print(f"States: {len(state_names)}")

# ── 2. Year columns from _Data header ──
ws_data = wb["_Data"]
year_indices = {}
for cell in next(ws_data.iter_rows(min_row=1, max_row=1)):
    v = cell.value
    if v is not None:
        try:
            yr = int(float(str(v)))
            if 1989 <= yr <= 2024:
                year_indices[cell.column] = yr
        except (ValueError, TypeError):
            pass
years = sorted(y for y in year_indices.values() if y >= 1990)
print(f"Years: {years[0]}-{years[-1]} ({len(years)} years)")

# ── 3. Raw variable codes to extract ──
raw_codes = [
    "NGSDP",
    # Current expenditure
    "EXP_CUR_TOTAL",
    "EXP_CUR_I",          # Current developmental
    "EXP_CUR_I_A",        # Social Services (current dev)
    "EXP_CUR_I_B",        # Economic Services (current dev)
    "EXP_CUR_II",         # Current non-developmental
    "EXP_CUR_II_A",       # Organs of State
    "EXP_CUR_II_B",       # Fiscal Services
    "EXP_CUR_II_C",       # Interest Payments and Servicing of Debt
    "EXP_CUR_II_D",       # Administrative Services
    "EXP_CUR_II_E",       # Pensions
    "EXP_CUR_II_F",       # Miscellaneous General Services
    "EXP_CUR_III_CAL",    # Compensation and Assignments to Local Bodies
    # Capital expenditure
    "EXP_CAP_I",          # Total Capital Outlay
    "EXP_CAP_I_1_DEV",    # Development Capital Outlay
    "EXP_CAP_I_1_A_SS",   # Social Services (capital dev)
    "EXP_CAP_I_1_B_ES",   # Economic Services (capital dev)
    "EXP_CAP_I_2_ND",     # Non-Development Capital Outlay
    "EXP_CAP_IV",         # Loans and Advances
    "EXP_CAP_IV_DEV",     # Development Purposes
    "EXP_CAP_IV_A_SS",    # Social Services (loans)
    "EXP_CAP_IV_B_ES",    # Economic Services (loans)
    "EXP_CAP_IV_2_ND",    # Non-Development Purposes
]

# Build lookup: key_pattern -> (state, code)
target_keys = {}
for state in state_names:
    for code in raw_codes:
        key = state + code + "None"
        target_keys[key] = (state, code)

# ── 4. Read raw data from _Data ──
raw_data = {s: {} for s in state_names}
for row in ws_data.iter_rows(min_row=2, values_only=False):
    key_cell = row[0].value
    if not key_cell:
        continue
    key = str(key_cell)
    if key in target_keys:
        state, code = target_keys[key]
        raw_data[state][code] = {}
        for col_idx, yr in year_indices.items():
            cell = row[col_idx - 1]
            if cell.value is not None:
                try:
                    raw_data[state][code][yr] = float(cell.value)
                except (ValueError, TypeError):
                    pass

found = sum(1 for s in state_names for c in raw_codes if c in raw_data[s])
print(f"Extracted {found} state-code combinations (expected {len(state_names) * len(raw_codes)})")

# ── 5. Compute derived metrics ──
output_rows = []

# Expenditure indicators: (raw_code, output_code, display_label)
exp_indicators = [
    ("EXP_CUR_TOTAL",    "cur_total",    "Total current expenditure"),
    ("EXP_CUR_I",        "cur_dev",      "Current developmental expenditure"),
    ("EXP_CUR_I_A",      "cur_dev_ss",   "Social Services"),
    ("EXP_CUR_I_B",      "cur_dev_es",   "Economic Services"),
    ("EXP_CUR_II",       "cur_nondev",   "Current non-developmental expenditure"),
    ("EXP_CUR_II_A",     "cur_nd_org",   "Organs of State"),
    ("EXP_CUR_II_B",     "cur_nd_fis",   "Fiscal Services"),
    ("EXP_CUR_II_C",     "cur_nd_int",   "Interest Payments and Servicing of Debt"),
    ("EXP_CUR_II_D",     "cur_nd_adm",   "Administrative Services"),
    ("EXP_CUR_II_E",     "cur_nd_pen",   "Pensions"),
    ("EXP_CUR_II_F",     "cur_nd_mis",   "Miscellaneous General Services"),
    ("EXP_CUR_III_CAL",  "cur_comp",     "Compensation and Assignments to Local Bodies and PRIs"),
    ("EXP_CAP_I",        "cap_outlay",   "Total Capital Outlay"),
    ("EXP_CAP_I_1_DEV",  "cap_dev",      "Development Capital Outlay"),
    ("EXP_CAP_I_1_A_SS", "cap_dev_ss",   "Social Services"),
    ("EXP_CAP_I_1_B_ES", "cap_dev_es",   "Economic Services"),
    ("EXP_CAP_I_2_ND",   "cap_nondev",   "Non-Development Capital Outlay"),
    ("EXP_CAP_IV",       "loans",        "Loans and Advances"),
    ("EXP_CAP_IV_DEV",   "loans_dev",    "Development Purposes"),
    ("EXP_CAP_IV_A_SS",  "loans_ss",     "Social Services"),
    ("EXP_CAP_IV_B_ES",  "loans_es",     "Economic Services"),
    ("EXP_CAP_IV_2_ND",  "loans_nondev", "Non-Development Purposes"),
]

for state in state_names:
    d = raw_data[state]
    for yr in years:
        ngdp = d.get("NGSDP", {}).get(yr)
        if not ngdp or ngdp == 0:
            continue

        # Total expenditure = current + capital outlay + loans
        cur_total = d.get("EXP_CUR_TOTAL", {}).get(yr, 0)
        cap_i     = d.get("EXP_CAP_I", {}).get(yr, 0)
        cap_iv    = d.get("EXP_CAP_IV", {}).get(yr, 0)
        total_exp = cur_total + cap_i + cap_iv

        # Total expenditure % of GDP
        output_rows.append([state, "total_exp_gdp", "Total", yr,
                            round(total_exp / ngdp * 100, 2)])

        # Each component as % of GDP
        for raw_code, out_code, label in exp_indicators:
            raw_val = d.get(raw_code, {}).get(yr)
            if raw_val is not None:
                pct_gdp = raw_val / ngdp * 100
                output_rows.append([state, f"{out_code}_gdp", label, yr,
                                    round(pct_gdp, 2)])

        # Percent of total expenditure (only if total > 0)
        if total_exp > 0:
            for raw_code, out_code, label in exp_indicators:
                raw_val = d.get(raw_code, {}).get(yr)
                if raw_val is not None:
                    pct_total = raw_val / total_exp * 100
                    output_rows.append([state, f"{out_code}_total", label, yr,
                                        round(pct_total, 2)])

# ── 6. Write CSV ──
with open(OUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["State", "Code", "Indicator", "Year", "Value"])
    writer.writerows(output_rows)

print(f"\nWritten {len(output_rows)} rows to {OUT}")

# Verify
for r in output_rows:
    if r[0] == "Bihar" and r[1] == "total_exp_gdp" and r[3] == 2017:
        print(f"Bihar total_exp_gdp 2017: {r[4]}")
    if r[0] == "Bihar" and r[1] == "cur_total_gdp" and r[3] == 2017:
        print(f"Bihar cur_total_gdp 2017: {r[4]}")
    if r[0] == "Bihar" and r[1] == "cap_outlay_gdp" and r[3] == 2017:
        print(f"Bihar cap_outlay_gdp 2017: {r[4]}")

from collections import Counter
code_counts = Counter(r[1] for r in output_rows)
print(f"\nTotal indicator codes: {len(code_counts)}")
for code, count in sorted(code_counts.items()):
    print(f"  {code}: {count} rows")

wb.close()
